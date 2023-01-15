import openpyxl
import datetime


symbols = ["A", "B", "C", "D", "E", "F", "G", "H", 'I', 'J', "K", 'L', 'M', 'N', 'O', 'P', 'Q', "R", 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
           'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG',
           'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY',
           'BZ']


def create_result_file(criteria, region):
    current_time = "-".join(str(datetime.datetime.now().time()).split(".")[:-1]).replace(":", "-")
    file_name = f"result/{datetime.date.today()}-{current_time}-{criteria}-{region}.xlsx"
    workbook = openpyxl.Workbook()
    page = workbook.active
    page[f"A1"].value = "Артикул"
    page[f"B1"].value = "SKU Товара"
    page[f"C1"].value = "Название"
    page[f"D1"].value = "Ссылка на товар"
    page[f"E1"].value = "Модель"
    page[f"F1"].value = "Цена со скидкой"
    page[f"G1"].value = "Цена без скидки"
    page[f"H1"].value = "Цена с картой Ozon"
    page[f"I1"].value = "Категории"
    page[f"J1"].value = "Главное изображение"
    page[f"K1"].value = "Изображения"
    page[f"L1"].value = "ID главного изображения"
    page[f"M1"].value = "Рейтинг"
    page[f"N1"].value = "Количество отзывов"
    page[f"O1"].value = "Продавец"
    page[f"P1"].value = "Описание"
    workbook.save(file_name)
    return file_name


def get_start_data():
    result = {"sku": list(), "product_link": list(), "brand": list(), "seller": list(), "search_request": list()}
    workbook = openpyxl.load_workbook("Входящие данные.xlsx")
    page = workbook.active
    for index in range(2, page.max_row + 1):
        sku = str(page[f"A{index}"].value).strip()
        if len(sku) > 0 and sku != "None":
            result["sku"].append({"value": sku})
        else:
            sku_list = str(page[f"B{index}"].value).strip()
            if len(sku_list) > 0 and sku_list != "None":
                sku_list = sku_list.split(",")
                for element in sku_list:
                    element = element.strip()
                    result["sku"].append({"value": element})
            else:
                product_link = str(page[f"C{index}"].value).strip()
                if len(product_link) > 0 and product_link != "None":
                    result["product_link"].append({"value": product_link})
                else:
                    product_links = str(page[f"D{index}"].value).strip()
                    if len(product_links) > 0 and product_links != "None":
                        product_links = product_links.split(",")
                        for element in product_links:
                            element = element.strip()
                            result["product_link"].append({"value": element})
                    else:
                        brands = str(page[f"E{index}"].value).strip()
                        if len(brands) > 0 and brands != "None":
                            brands = brands.split(",")
                            for element in brands:
                                element = element.strip()
                                result["brand"].append({"value": element})
                        else:
                            seller = str(page[f"F{index}"].value).strip()
                            if len(seller) > 0 and seller != "None":
                                result["seller"].append({"value": seller})
                            else:
                                search_request = str(page[f"G{index}"].value).strip()
                                if len(search_request) > 0 and search_request != "None":
                                    result["search_request"].append({"value": search_request})
    return result


def record_no_data(ozon_id, file_name, message):
    workbook = openpyxl.load_workbook(file_name)
    page = workbook.active
    index = page.max_row + 1
    page[f"B{index}"].value = ozon_id
    page[f"C{index}"].value = message
    workbook.save(file_name)


def record_data(article, ozon_id, product_name, model_name, purchase_price, full_price, discount_card_price,
                categories, main_image, additional_images, main_image_id, characteristics, rating, amount_reviews,
                file_name, seller, description, product_url):
    workbook = openpyxl.load_workbook(file_name)
    page = workbook.active
    index = page.max_row + 1
    page[f"A{index}"].value = article
    page[f"B{index}"].value = ozon_id
    page[f"C{index}"].value = product_name
    page[f"D{index}"].value = product_url
    page[f"E{index}"].value = model_name
    page[f"F{index}"].value = purchase_price
    page[f"G{index}"].value = full_price
    page[f"H{index}"].value = discount_card_price
    page[f"I{index}"].value = categories
    page[f"J{index}"].value = main_image
    page[f"K{index}"].value = additional_images
    page[f"L{index}"].value = main_image_id
    page[f"M{index}"].value = rating
    page[f"N{index}"].value = amount_reviews
    page[f"O{index}"].value = seller
    page[f"P{index}"].value = description
    for key in characteristics.keys():
        for symbol in symbols:
            current_header = page[f"{symbol}1"].value
            if current_header is not None:
                if key == current_header:
                    page[f"{symbol}{index}"].value = characteristics[key]
                    break
            else:
                page[f"{symbol}1"].value = key
                page[f"{symbol}{index}"].value = characteristics[key]
                break
    workbook.save(file_name)
    print(f"[INFO] Данные о товаре с Ozon ID = {ozon_id} были успешно записаны в таблицу")


if __name__ == "__main__":
    create_result_file(criteria="test", region="Москва")
